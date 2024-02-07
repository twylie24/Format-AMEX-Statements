import openpyxl
import os
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
import warnings
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# this script will format a pre-existing excel file to fit specific needs of travel billing: working with 'Transaction Details' sheet


# ignores the worning that workbook does not have 'Default' style
warnings.simplefilter("ignore", UserWarning)

# directory to find the excel file that we want to work with
# type name of file you are working with in name variable
name = input('File Name: ')
extension = '.xlsx'
file_name = f'{name}{extension}'
# directory that the file is in
dir = f"C:\\Users\\twyli\\OneDrive\\Desktop\\PSS\\Scripts\\Format AMEX Sheets\\{file_name}"


# load excel workbook: change file name when using different workbook - will start with example workbook
try:
    workbook = openpyxl.load_workbook(dir)
    print(f"Workbook loaded successfully: {dir}")
except FileNotFoundError:
    print(f"Error: File not found at {dir}")
except Exception as e:
    print(f"Error loading workbook: {e}")

# choose the sheet you want to work with - typically working with sheet named 'Transaction Details'
worksheet = workbook['Transaction Details']

# delete rows that are not needed for the work - will typically delete rows 1-6
rows_to_delete = [1, 2, 3, 4, 5, 6]

# loop over rows in 'delete_rows' set and delete them (reverse=True so that it deletes in backwards fashion so that excel row indicies stay the same while looping over set)
for row in sorted(rows_to_delete, reverse=True):
    worksheet.delete_rows(row)
    
# delete columns that are not needed for the sheet - typically delete columns B, D, E, H, I, J, K, L, M, N (convert to numbered columns)
columns_to_delete = [2, 4, 5, 8, 9, 10, 11, 12, 13, 14]

# loop over columns in 'delete_columns' set and delete them (reverse=True so that it deletes in backwards fashion so that excel row indicies stay the same while looping over set)
for column in sorted(columns_to_delete, reverse=True):
    worksheet.delete_cols(column)
    

# create a new column in between columns B & C titled 'Customer Class Codes'
class_codes_column = 'Customer Class Code'
codes_column_index = 4
worksheet.insert_cols(codes_column_index)
codes_header_cell = worksheet.cell(row=1, column=codes_column_index, value=class_codes_column)

# creating the rest of new columns needed - 6: 'Dates Covered By Charge', 7: 'FG', 8: 'Notes'
# 'Dates Covered By Charge'
dates_covered_column = 'Date Covered By Charge'
dates_covered_index = 6
worksheet.insert_cols(dates_covered_index)
dates_header_cell = worksheet.cell(row=1, column=dates_covered_index, value=dates_covered_column)

# 'FG'
fg_column = 'FG'
fg_index = 7
worksheet.insert_cols(fg_index)
fg_header_cell = worksheet.cell(row=1, column=fg_index, value=fg_column)

# 'Notes'
notes_column = 'Notes'
notes_index = 8
worksheet.insert_cols(notes_index)
notes_header_cell = worksheet.cell(row=1, column=notes_index, value=notes_column)

# formatting all the rows to have the new/correct styles (font, alignment, and highlight)
font_text = Font(color='000000', bold=True)
alignment_text = Alignment(horizontal='center')
fill_text = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')

# change name of column E to 'Employee Name'
new_name_E = 'Employee Name'
worksheet.cell(row=1, column=5).value = new_name_E

# loops over column headers and changes their font to bold, text to light gray, alignment to centered, and highlight to light blue
rows_to_format = [1]
columns_to_format = [1, 2, 3, 4, 5, 6, 7, 8]
for column_index in columns_to_format:
    cell_to_format = worksheet.cell(row=1, column=column_index)
    cell_to_format.font = font_text
    cell_to_format.alignment = alignment_text
    cell_to_format.fill = fill_text


last_row = worksheet.max_row
while last_row > 1 and not worksheet.cell(row=last_row, column=1).value:
    last_row -= 1

# specify the range for the border
border_range = f'A2:H{last_row}'

# create a border with inside and outside lines in plain black
border = Border(
    left=Side(style='thin', color='000000'),    # Left border
    right=Side(style='thin', color='000000'),   # Right border
    top=Side(style='thin', color='000000'),     # Top border
    bottom=Side(style='thin', color='000000'),  # Bottom border
)

# apply the border to the specified range
for row in worksheet.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=8):
    for cell in row:
        cell.border = border

# Set row height for all rows within the specified range
for row_index in range(2, last_row + 1):
    worksheet.row_dimensions[row_index].height = 45
    
    
# set fill color for all cells within the specified range to white
fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
for row in worksheet.iter_rows(min_row=2, max_row=last_row, min_col=1, max_col=8):
    for cell in row:
        cell.fill = fill_white

# center all of the cells for columns 3, 4, 6, 7 from row 2 to last row
columns_to_center = [3, 4, 6, 7]
for col_index in columns_to_center:
    for row_index in range(2, last_row + 1):
        cell_to_center = worksheet.cell(row=row_index, column=col_index)
        cell_to_center.alignment = Alignment(horizontal='center')


# make it so that the customer class code column will only accept whole numbers - no decimal points
validation_range = f'C2:C{last_row}'
data_validation = DataValidation(
    type="whole",
    operator=None,
    formula1=None,
    formula2=None,
    showErrorMessage=True,
    errorTitle="Invalid Input",
    error="Please enter a whole number.",
)
worksheet.add_data_validation(data_validation)
data_validation.add(validation_range)


# add a filter to all columns 1-8
worksheet.auto_filter.ref = worksheet.dimensions

# reduce the zoom to 85%
worksheet.sheet_view.zoomScale = 85




# save the modified workbook
workbook.save(dir)





